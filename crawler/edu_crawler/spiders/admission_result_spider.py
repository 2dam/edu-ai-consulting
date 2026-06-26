"""
대학 입학처 "입시 결과분석" 게시글에서 첨부된 엑셀(xlsx) 파일을 내려받아
학과별 경쟁률·입결(cutoff) 데이터를 추출하는 스파이더.

대학 입시결과는 schoolinfo.go.kr(학교알리미)나 NEIS Open API에는 없고,
각 대학 입학처 홈페이지의 공고 게시글에 PDF/엑셀 첨부파일로만 공시된다.
이 스파이더는 그 게시글 상세 페이지를 받아 첨부파일 다운로드 링크를 찾고,
다운로드된 엑셀을 openpyxl로 직접 파싱한다 (HTML 표 셀렉터가 아님).

대학별 robots.txt 점검 결과 (2026-06-26 확인):
  허용 — 고려대, 부산대, 서울대, 연세대 (일반 User-agent 차단 규칙 없음)
  차단 — 전남대, 충남대, 경북대 (User-agent: * Disallow: / 명시) → 절대 크롤링하지 말 것

차단된 대학은 ROBOTSTXT_WHITELIST에 넣지 않는다. 화이트리스트에 없는 대학으로
실행을 시도하면 즉시 에러를 낸다 (robots.txt 미들웨어에 맡기지 않고 코드 차원에서도 한 번 더 막음).

실행 예 (전남대 2025학년도 수시 입시결과 게시글 — 데모용, 실제로는 차단되어 동작 안 함):
  scrapy crawl admission_result \
    -a detail_url="https://admission.jnu.ac.kr/WebApp/web/HOM/COM/Board/board.aspx?boardID=423&bbsMode=view&page=1&key=25" \
    -a university="전남대학교" -a year="2025"

주의: 엑셀 시트의 헤더 행 수·컬럼 순서는 대학마다 다르므로, 새 대학을 추가할 때는
UNIVERSITY_COLUMN_MAP에 그 대학 파일에 맞는 컬럼 인덱스를 추가해야 한다.
"""
from datetime import datetime, timezone
from io import BytesIO

import openpyxl
import scrapy

from edu_crawler.items import AdmissionResultItem

# robots.txt가 일반 크롤러를 허용하는 대학만 등재 (2026-06-26 점검 기준).
# 전남대/충남대/경북대는 User-agent: * Disallow: / 이므로 절대 추가하지 말 것.
ROBOTSTXT_WHITELIST = {"고려대학교", "부산대학교", "서울대학교", "연세대학교"}

# 대학별 엑셀 헤더 행 수 / 컬럼 인덱스 매핑.
# 전남대만 실측 확인됨 (화이트리스트 외 대학이라 참고용 기본값으로만 사용).
UNIVERSITY_COLUMN_MAP = {
    "전남대학교": {
        "header_rows": 7,
        "college": 3,
        "department": 4,
        "admission_type": 2,
        "competition_rate": 8,
        "cutoff_70": 18,
    },
}


class AdmissionResultSpider(scrapy.Spider):
    name = "admission_result"
    custom_settings = {"ROBOTSTXT_OBEY": True}

    def __init__(self, detail_url=None, university=None, year=None, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if not detail_url or not university:
            raise ValueError(
                "usage: -a detail_url=<입시결과 게시글 URL> -a university=<대학명> [-a year=<연도>]"
            )
        if university not in ROBOTSTXT_WHITELIST:
            raise ValueError(
                f"'{university}'는 robots.txt 화이트리스트에 없습니다. "
                f"먼저 해당 대학 robots.txt를 점검하고 ROBOTSTXT_WHITELIST에 추가하세요. "
                f"허용된 대학: {sorted(ROBOTSTXT_WHITELIST)}"
            )
        if university not in UNIVERSITY_COLUMN_MAP:
            raise ValueError(
                f"'{university}'는 화이트리스트에는 있지만 엑셀 컬럼 매핑이 아직 없습니다. "
                f"해당 대학 입시결과 엑셀을 받아 UNIVERSITY_COLUMN_MAP에 매핑을 추가하세요."
            )
        self.start_urls = [detail_url]
        self.university = university
        self.year = year or ""
        self.column_map = UNIVERSITY_COLUMN_MAP[university]

    def parse(self, response):
        for href in response.css('a[href*="bbsMode=download"]::attr(href)').getall():
            file_url = response.urljoin(href)
            yield scrapy.Request(
                file_url,
                callback=self.parse_excel,
                headers={"Referer": response.url},
            )

    def parse_excel(self, response):
        now = datetime.now(timezone.utc).isoformat()
        wb = openpyxl.load_workbook(BytesIO(response.body), data_only=True)
        ws = wb[wb.sheetnames[0]]
        cm = self.column_map

        for row in ws.iter_rows(min_row=cm["header_rows"] + 1, values_only=True):
            department_cell = row[cm["department"]] if len(row) > cm["department"] else None
            if not department_cell:
                continue

            college = row[cm["college"]] or ""
            item = AdmissionResultItem()
            item["source_url"] = response.url
            item["university"] = self.university
            item["department"] = f"{college} {department_cell}".strip()
            item["year"] = self.year
            item["admission_type"] = row[cm["admission_type"]] or ""
            item["competition_rate"] = row[cm["competition_rate"]]
            item["cutoff_score"] = row[cm["cutoff_70"]]
            item["crawled_at"] = now
            yield item
